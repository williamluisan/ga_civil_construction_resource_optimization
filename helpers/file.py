from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil

def OpenPyxlReadExcel(filename: str, data_only: bool=True) -> any:
    wb = load_workbook(filename, data_only = data_only)
    return wb

def copy_file(file_source: str, file_destination: str) -> any:
    try:
        shutil.copyfile(file_source, file_destination)
    except IOError as e:
        print(f"An error occurred while copying the file: {e}")
        return False
    
    return file_destination